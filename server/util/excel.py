from server.dataset import Data
from typing import List, Dict
from openpyxl import Workbook
from datetime import datetime

import os
import numpy as np


class ExcelUtil:

    CORAL_H = "Coral"
    CORAL_ID_H = "Coral ID"
    NUM_OF_PIXELS_H = "No. of Pixels"
    NUM_OF_HEALTHY_PIXEL_H = "No. of Healthy Pixel"
    NUM_OF_BLEACHED_PIXEL_H = "No. of Bleached Pixel"
    CORAL_COVERAGE_H = "Coral Coverage"
    HEALTHY_COVERAGE_H = "Healthy Coverage"
    BLEACHED_COVERAGE_H = "Bleached Coverage"
    HEALTHY_DISTRIBUTION_H = "Healthy Distribution"
    BLEACHED_DISTRIBUTION_H = "Bleached Distribution"
    NUM_OF_COLONY_H = "No. of Colony"

    def __init__(self, category_info: List[Dict]):
        """
        Input is the list of category info
        {
            "id": -1,
            "name": "Detected Coral",
            "supercategory": "Detected Coral",
            "supercategory_id": -1,
            "is_coral": True,
            "status": -1,
        }

        """
        self.category_dict = {}
        for category in category_info:
            self.category_dict[int(category["id"])] = category

    def export_excel(self, data: Data, output_path: str):

        wb = Workbook()
        ws = wb.active
        filename = os.path.basename(data.get_image_path())
        ws.title = filename

        ws["A1"] = "Image Name"
        ws["B1"] = filename

        image_width = data.get_image_width()
        image_height = data.get_image_height()
        ws["A2"] = "Image Pixel"
        ws["B2"] = image_width * image_height

        ws["A3"] = "Export Date"
        ws["B3"] = datetime.now().strftime("%d/%m/%Y")

        headers = [
            ExcelUtil.CORAL_H,
            ExcelUtil.CORAL_ID_H,
            ExcelUtil.NUM_OF_PIXELS_H,
            ExcelUtil.NUM_OF_HEALTHY_PIXEL_H,
            ExcelUtil.NUM_OF_BLEACHED_PIXEL_H,
            ExcelUtil.CORAL_COVERAGE_H,
            ExcelUtil.HEALTHY_COVERAGE_H,
            ExcelUtil.BLEACHED_COVERAGE_H,
            ExcelUtil.HEALTHY_DISTRIBUTION_H,
            ExcelUtil.BLEACHED_DISTRIBUTION_H,
            ExcelUtil.NUM_OF_COLONY_H,
        ]

        ws.append([])
        ws.append(headers)

        excel_data = self.extract_excel_data(data)
        excel_data = dict(sorted(excel_data.items()))

        for _, value in excel_data.items():
            row = [value[header] for header in headers]
            ws.append(row)

        ws.append([])

        # Calculate the sum of each column
        sum_info_list = []
        for header in headers:
            if header == ExcelUtil.CORAL_H:
                sum_info_list.append("")
                continue

            if header == ExcelUtil.CORAL_ID_H:
                sum_info_list.append("Sum")
                continue

            values = [value[header] for value in excel_data.values()]
            sum_info_list.append(sum(values))
        ws.append(sum_info_list)

        # Calculate the average of each column
        average_info_list = []
        for header in headers:
            if header == ExcelUtil.CORAL_H:
                average_info_list.append("")
                continue

            if header == ExcelUtil.CORAL_ID_H:
                average_info_list.append("Average")
                continue

            values = [value[header] for value in excel_data.values()]
            if len(values) == 0:
                average_info_list.append(0)
            else:
                average_info_list.append(sum(values) / len(values))
        ws.append(average_info_list)

        # Calculate the standard deviation of each column
        std_info_list = []
        for header in headers:
            if header == ExcelUtil.CORAL_H:
                std_info_list.append("")
                continue

            if header == ExcelUtil.CORAL_ID_H:
                std_info_list.append("Standard Deviation")
                continue

            values = [value[header] for value in excel_data.values()]
            if len(values) == 0:
                std_info_list.append(0)
            else:
                std_info_list.append(np.std(values))
        ws.append(std_info_list)

        ws.append([ExcelUtil.CORAL_H, "The name of the coral genus"])
        ws.append([ExcelUtil.CORAL_ID_H, "The ID of the coral genus"])
        ws.append(
            [ExcelUtil.NUM_OF_PIXELS_H, "The number of pixels of the coral genus"]
        )
        ws.append(
            [
                ExcelUtil.NUM_OF_HEALTHY_PIXEL_H,
                "The number of healthy pixels of the coral genus",
            ]
        )
        ws.append(
            [
                ExcelUtil.NUM_OF_BLEACHED_PIXEL_H,
                "The number of bleached pixels of the coral genus",
            ]
        )
        ws.append(
            [
                ExcelUtil.CORAL_COVERAGE_H,
                "The coverage of the coral genus: number of pixels / image pixel",
            ]
        )
        ws.append(
            [
                ExcelUtil.HEALTHY_COVERAGE_H,
                "The coverage of the healthy coral genus: number of healthy pixels / image pixel",
            ]
        )
        ws.append(
            [
                ExcelUtil.BLEACHED_COVERAGE_H,
                "The coverage of the bleached coral genus: number of bleached pixels / image pixel",
            ]
        )
        ws.append(
            [
                ExcelUtil.HEALTHY_DISTRIBUTION_H,
                "The distribution of the healthy coral genus: number of healthy pixels / number of coral pixels",
            ]
        )
        ws.append(
            [
                ExcelUtil.BLEACHED_DISTRIBUTION_H,
                "The distribution of the bleached coral genus: number of bleached pixels / number of coral pixels",
            ]
        )
        ws.append([ExcelUtil.NUM_OF_COLONY_H, "The number of coral colony"])

        wb.save(output_path)

    def extract_excel_data(self, data: Data):
        excel_data = {}

        image_width = data.get_image_width()
        image_height = data.get_image_height()
        image_pixel = image_width * image_height

        segmentation = data.get_segmentation()
        for annotation in segmentation["annotations"]:
            category_id = annotation["category_id"]
            category_info = self.category_dict[int(category_id)]

            super_category_name = category_info["supercategory"]
            super_category_id = category_info["supercategory_id"]
            super_category_id = int(super_category_id)

            status_id = category_info["status"]

            # We ignore the undefined category
            if status_id == Data.STATUS_UNDEFINED:
                continue

            is_coral = category_info["is_coral"]
            is_coral = bool(is_coral)

            # Now we only consider the coral category
            if not is_coral:
                continue

            if super_category_id not in excel_data:
                excel_data[super_category_id] = {}
                excel_data[super_category_id][ExcelUtil.CORAL_H] = super_category_name
                excel_data[super_category_id][ExcelUtil.CORAL_ID_H] = super_category_id
                excel_data[super_category_id][ExcelUtil.NUM_OF_PIXELS_H] = 0
                excel_data[super_category_id][ExcelUtil.NUM_OF_HEALTHY_PIXEL_H] = 0
                excel_data[super_category_id][ExcelUtil.NUM_OF_BLEACHED_PIXEL_H] = 0
                excel_data[super_category_id][ExcelUtil.CORAL_COVERAGE_H] = 0
                excel_data[super_category_id][ExcelUtil.HEALTHY_COVERAGE_H] = 0
                excel_data[super_category_id][ExcelUtil.BLEACHED_COVERAGE_H] = 0
                excel_data[super_category_id][ExcelUtil.HEALTHY_DISTRIBUTION_H] = 0
                excel_data[super_category_id][ExcelUtil.BLEACHED_DISTRIBUTION_H] = 0
                excel_data[super_category_id][ExcelUtil.NUM_OF_COLONY_H] = 0

            if status_id == Data.STATUS_BLEACHED:
                excel_data[super_category_id][
                    ExcelUtil.NUM_OF_BLEACHED_PIXEL_H
                ] += annotation["area"]
            else:
                excel_data[super_category_id][
                    ExcelUtil.NUM_OF_HEALTHY_PIXEL_H
                ] += annotation["area"]

            excel_data[super_category_id][ExcelUtil.NUM_OF_PIXELS_H] += annotation[
                "area"
            ]
            excel_data[super_category_id][ExcelUtil.NUM_OF_COLONY_H] += 1

        for super_category_id, info in excel_data.items():
            info[ExcelUtil.CORAL_COVERAGE_H] = (
                info[ExcelUtil.NUM_OF_PIXELS_H] / image_pixel
            )
            info[ExcelUtil.HEALTHY_COVERAGE_H] = (
                info[ExcelUtil.NUM_OF_HEALTHY_PIXEL_H] / image_pixel
            )
            info[ExcelUtil.BLEACHED_COVERAGE_H] = (
                info[ExcelUtil.NUM_OF_BLEACHED_PIXEL_H] / image_pixel
            )

            if info[ExcelUtil.NUM_OF_PIXELS_H] == 0:
                info[ExcelUtil.HEALTHY_DISTRIBUTION_H] = 0
                info[ExcelUtil.BLEACHED_DISTRIBUTION_H] = 0
            else:
                info[ExcelUtil.HEALTHY_DISTRIBUTION_H] = (
                    info[ExcelUtil.NUM_OF_HEALTHY_PIXEL_H]
                    / info[ExcelUtil.NUM_OF_PIXELS_H]
                )
                info[ExcelUtil.BLEACHED_DISTRIBUTION_H] = (
                    info[ExcelUtil.NUM_OF_BLEACHED_PIXEL_H]
                    / info[ExcelUtil.NUM_OF_PIXELS_H]
                )

        return excel_data
